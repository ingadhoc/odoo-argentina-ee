import base64
import datetime
import io
import logging
import numbers

import xlrd
from odoo import api, fields, models
from odoo.exceptions import ValidationError
from openpyxl import load_workbook

_logger = logging.getLogger(__name__)


class AccountBalanceImport(models.TransientModel):
    _name = "account.balance_import_wizard"
    _description = "Account Initial Balance Wizard"
    _check_company_auto = True
    _check_company_domain = models.check_companies_domain_parent_of

    # Common Fields
    company_id = fields.Many2one(
        "res.company",
        string="Compañía",
        required=True,
        default=lambda self: self.env.company,
    )

    counterpart_account_id = fields.Many2one(
        "account.account",
        string="Cuenta de Contrapartida",
        default=lambda self: self.env.company.get_unaffected_earnings_account(),
        help="Recomendamos utilizar la misma cuenta de contrapartida para todos los asientos iniciales",
        check_company=True,
    )

    mode = fields.Selection(
        [
            ("partner_balance", "Saldos de Partners"),
        ],
        string="Modo de Importación",
        required=True,
        default="partner_balance",
        ondelete={"partner_balance": "set default"},
    )
    accounting_date = fields.Date(
        "Fecha Contable", required=True, compute="_compute_accounting_date", store=True, readonly=False
    )

    # Company Balance Related Fields
    journal_id = fields.Many2one(
        "account.journal",
        string="Diario",
        domain="journal_domain",
        check_company=True,
    )
    journal_domain = fields.Binary(
        compute="_compute_journal_domain",
    )
    partner_balance_type = fields.Selection(
        [("receivable", "Por Cobrar"), ("payable", "A Pagar")],
        "Tipo de Deuda",
        default="receivable",
    )
    export_partners = fields.Boolean(
        string="Exportar Partners",
        default=True,
        help="Si se activa, la plantilla incluirá los clientes o proveedores existentes según el Tipo de Deuda seleccionado.",
    )
    import_type = fields.Selection(
        [
            ("absolute", "Absoluto"),
            ("adjust", "Ajuste"),
        ],
        string="Tipo de Importación",
        default="absolute",
        required=True,
        help="Absoluto: genera asientos directamente con los valores importados.\n"
        "Ajuste: calcula la diferencia entre el saldo actual del partner y el valor importado.",
    )
    reconcile_debt = fields.Boolean(
        string="Conciliar deuda",
        default=True,
        help="si marca esta opción se conciliará automáticamente con la deuda más vieja.",
    )
    file = fields.Binary("Archivo de Importación")

    @api.depends("mode")
    def _compute_journal_domain(self):
        for rec in self:
            if rec.mode == "partner_balance":
                rec.journal_domain = [("type", "=", "general")]
            else:
                rec.journal_domain = []

    @api.onchange("company_id")
    def _onchange_company_id(self):
        """Update counterpart account when company changes"""
        if self.company_id:
            self.counterpart_account_id = self.company_id.get_unaffected_earnings_account()

    @api.model
    def default_get(self, fields):
        res = super(AccountBalanceImport, self).default_get(fields)

        # Verificamos si el contexto tiene un valor para 'default_mode'
        if "default_mode" in self.env.context:
            res["mode"] = self.env.context["default_mode"]

        return res

    @api.depends("company_id.account_opening_date")
    def _compute_accounting_date(self):
        self.accounting_date = (
            self.company_id.account_opening_date or fields.Date.start_of(fields.Date.today(), "year")
        ) - datetime.timedelta(days=1)

    def action_generate_partner_balance_xls(self):
        self.ensure_one()
        params = {
            "company_id": self.company_id.id,
            "date": self.accounting_date.isoformat() if self.accounting_date else "",
            "balance_type": self.partner_balance_type or "receivable",
            "export_partners": "1" if self.export_partners else "",
            "import_type": self.import_type or "absolute",
        }
        url = "/account_balance_import/partner_balance_xls?" + "&".join(f"{k}={v}" for k, v in params.items())
        return {
            "type": "ir.actions.act_url",
            "url": url,
            "target": "new",
        }

    @api.model
    def locate_currency(self, currency):
        """This method return the currency, if wasn't find any currency that match with the name in xls we return empty recordset"""
        other_currency = self.env["res.currency"]
        if currency:
            other_currency = self.env["res.currency"]._search_by_name(currency)
        return other_currency

    def action_import(self):
        self.ensure_one()
        if not self.file:
            raise ValidationError("Por favor, cargue un archivo Excel para importar.")
        if self.mode == "partner_balance":
            return self._partner_balance_import_xls()

    @api.model
    def _get_partners_balances_at_date(self, partners, date_to, balance_type, company):
        """Get balances for multiple partners at a specific date.

        Uses _read_group for better performance when processing many partners.

        Args:
            partners: res.partner recordset
            date_to: date to calculate balances up to
            balance_type: 'receivable' or 'payable'
            company: res.company record

        Returns:
            dict: {partner_id: balance} where balance is positive for amount owed
        """
        if not partners:
            return {}

        account_type = "asset_receivable" if balance_type == "receivable" else "liability_payable"
        domain = [
            ("partner_id", "in", partners.ids),
            ("date", "<=", date_to),
            ("parent_state", "=", "posted"),
            ("account_id.account_type", "=", account_type),
            ("company_id", "=", company.id),
        ]

        # Use _read_group for efficient aggregation
        # balance:sum returns debit - credit
        results = self.env["account.move.line"]._read_group(
            domain,
            ["partner_id"],
            ["balance:sum"],
        )

        # Build result dict
        balances = {}
        for partner, balance_sum in results:
            # For receivable: positive balance means customer owes us
            # For payable: flip sign so positive means we owe the vendor
            if balance_type == "payable":
                balance_sum = -balance_sum
            balances[partner.id] = balance_sum

        return balances

    def _get_partner_balance(self, partner, balance_type):
        """Get the current balance for a single partner up to the accounting date.

        Args:
            partner: res.partner record
            balance_type: 'receivable' or 'payable'

        Returns:
            float: The partner's balance (positive = amount owed)
        """
        balances = self._get_partners_balances_at_date(partner, self.accounting_date, balance_type, self.company_id)
        return balances.get(partner.id, 0.0)

    def _partner_balance_import_xls(self):  # noqa: C901
        """Triggered on when Company Balance XLS is imported
        This function will firstly read the entire XLS file and perform
        checks on each one of the rows to make sure the import looks good.
        If everything goes fine, the account moves are created and displayed.
        If any error is found, a message with the offending lines and their
        description is presented.
        """

        # Set-up environment company
        self = self.with_company(self.company_id.id)
        fields = ["name", "reference", "amount"]
        if self.import_type != "adjust":
            fields.extend(
                [
                    "due_date",
                    "currency",
                    "amount_company_currency",
                ]
            )

        account_moves = list()  # For storing account moves before creating 'em
        errors = list()  # For storing possible errors
        company = self.env.company
        journal = self.journal_id

        # Parse XLS or XLSX file
        decoded = base64.decodebytes(self.file)
        sheet_rows = []
        try:
            # Try XLS format
            workbook = xlrd.open_workbook(file_contents=decoded)
            sheet = workbook.sheet_by_index(0)
            self._validate_column_count(sheet, fields)
            for row_no in range(1, sheet.nrows):
                values = [sheet.row(row_no)[i].value for i in range(len(fields))]
                sheet_rows.append((row_no + 1, values))
        except xlrd.biffh.XLRDError:
            # Fallback to XLSX format
            workbook = load_workbook(io.BytesIO(decoded), data_only=True)
            sheet = workbook.active
            for row_no, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                values = list(row[: len(fields)])
                sheet_rows.append((row_no, values))

        # Iterate over each sheet row
        for row_no, values in sheet_rows:
            dict_data = {fields[i]: values[i] for i in range(0, len(fields))}

            # Parse name as string to prevent CUIT being read as float
            if isinstance(dict_data["name"], numbers.Number):
                dict_data["name"] = str(int(dict_data["name"]))

            # Locate Partner
            domain = [
                "|",
                "|",
                ("name", "=", dict_data["name"]),
                ("vat", "=", dict_data["name"]),
                ("ref", "=", dict_data["name"]),
            ]

            partner = self.env["res.partner"].search(domain)

            # Locate Currency
            other_currency = self.locate_currency(dict_data.get("currency"))

            if other_currency and not dict_data.get("amount_company_currency"):
                errors.append(
                    f"Fila {str(row_no)}: Si le establece otra moneda debe indicar el importe en esa otra moneda"
                )
                continue

            # Skip if partner not found
            if not partner:
                errors.append(
                    f"Fila {str(row_no)}: No se encontró ningún partner para el texto ingresado ({dict_data['name']})."
                )
                continue

            # Skip if more than one partner was found
            if len(partner) > 1:
                errors.append(
                    f"Fila {str(row_no)}: Se encontraron varios partners para el texto ingresado ({dict_data['name']}). ¡Revise los datos cargados!"
                )
                continue

            # Skip if amount isn't numerical
            if not isinstance(dict_data["amount"], numbers.Number):
                errors.append(f"Fila {str(row_no)}: El monto no es numérico.")
                continue

            # Parse due_date
            if dict_data.get("due_date") and dict_data.get("due_date") != "":
                try:
                    if isinstance(dict_data["due_date"], datetime.date):
                        due_date = dict_data["due_date"]
                    else:
                        due_date = datetime.datetime(
                            *xlrd.xldate_as_tuple(dict_data["due_date"], getattr(workbook, "datemode", 0))
                        ).date()
                except Exception:
                    errors.append(
                        f"Fila {str(row_no)}: Formato de fecha de vencimiento desconocido. Asegúrese de que la columna posee formato de fecha."
                    )
                    continue
            else:
                due_date = False

            # Get account depending on selected type
            if self.partner_balance_type == "receivable":
                partner_account = partner.property_account_receivable_id
            else:
                partner_account = partner.property_account_payable_id

            # Calculate the amount to import based on import_type
            imported_amount = dict_data["amount"]
            if self.import_type == "adjust":
                # Get current partner balance using native credit/debit fields
                # These fields already return positive values for the amount owed
                current_balance = self._get_partner_balance(partner, self.partner_balance_type)
                # Calculate the difference needed to reach the target balance
                imported_amount = imported_amount - current_balance

            # Determine debit/credit based on balance type and amount
            if self.partner_balance_type == "receivable":
                if imported_amount > 0:
                    debit = self.company_id.currency_id.round(imported_amount)
                    credit = 0.0
                elif imported_amount < 0:
                    debit = 0.0
                    credit = self.company_id.currency_id.round(abs(imported_amount))
                else:
                    continue
            else:
                if imported_amount > 0:
                    debit = 0.0
                    credit = self.company_id.currency_id.round(imported_amount)
                elif imported_amount < 0:
                    debit = self.company_id.currency_id.round(abs(imported_amount))
                    credit = 0.0
                else:
                    continue

            # Skip if partner account account is obsolete
            if not partner_account.active:
                errors.append(
                    f"Fila {str(row_no)}: La cuenta asociada al partner {partner.name} se encuentra depreciada."
                )
                continue

            # Check if accounts have the same company
            if company.id not in partner_account.company_ids.ids:
                errors.append(
                    f"Fila {str(row_no)}: Una de las cuentas asociadas al partner {partner.name} no pertenece a la compañía ({company.name})"
                )
                continue

            amount_company_currency = (
                other_currency and abs(other_currency.round(dict_data["amount_company_currency"])) or False
            )

            # Create move lines
            line_1 = {
                "name": dict_data["reference"],
                "partner_id": partner.id,
                "account_id": partner_account.id,
                "debit": debit,
                "credit": credit,
                "date_maturity": due_date,
            }

            line_2 = {
                "name": dict_data["reference"],
                "partner_id": partner.id,
                "account_id": self.counterpart_account_id.id,
                "debit": credit,
                "credit": debit,
                "date_maturity": due_date,
            }

            if other_currency:
                line_1.update(
                    {
                        "currency_id": other_currency.id,
                        "amount_currency": (-1.0 if line_1["debit"] == 0.0 else 1.0) * abs(amount_company_currency),
                    }
                )
                line_2.update(
                    {
                        "currency_id": other_currency.id,
                        "amount_currency": (1.0 if line_2["credit"] == 0.0 else -1.0) * abs(amount_company_currency),
                    }
                )

            # Add account move to list
            account_moves.append(
                {
                    "journal_id": journal.id,
                    "ref": dict_data["reference"],
                    "date": self.accounting_date,
                    "invoice_date_due": due_date or self.accounting_date,
                    "line_ids": [
                        (0, 0, line_1),
                        (0, 0, line_2),
                    ],
                }
            )

        # Check if there were errors when iterating over XLS rows
        if len(errors) > 0:
            raise ValidationError("\n".join(errors))

        # Everything should be OK if we reached this part
        generated_moves = self.env["account.move"].create(account_moves)
        # Post Account Move
        generated_moves._post()

        if self.import_type == "adjust" and self.reconcile_debt:
            for move in generated_moves:
                # Find the line that corresponds to the partner (asset_receivable or liability_payable)
                partner_line = move.line_ids.filtered(
                    lambda l: l.account_id.account_type in ("asset_receivable", "liability_payable")
                )
                if not partner_line:
                    continue

                # Search for other non-reconciled items for the same partner and account
                # to reconcile with the oldest first
                domain = [
                    ("parent_state", "=", "posted"),
                    ("partner_id", "=", partner_line.partner_id.id),
                    ("account_id", "=", partner_line.account_id.id),
                    ("reconciled", "=", False),
                    ("id", "!=", partner_line.id),
                    ("date", "<=", self.accounting_date),
                ]
                # We can only reconcile if there are lines with opposite balance
                # Note: if it's an adjustment, either we have a debit to increase debt or a credit to decrease it.
                if partner_line.debit > 0:
                    domain.append(("credit", ">", 0))
                else:
                    domain.append(("debit", ">", 0))

                outstanding_lines = self.env["account.move.line"].search(domain, order="date asc, id asc")
                if outstanding_lines:
                    (partner_line | outstanding_lines).reconcile()

        # mandamos views para que tome la vista de menu asientos, si no por defecto toma follow up o alguna otra, hay varias primary con 16 de secuencia
        return generated_moves._get_records_action(
            views=[(self.env.ref("account.view_move_tree").id, "list")],
            domain=[("id", "in", generated_moves.ids)],
        )

    def _validate_column_count(self, sheet, expected_fields):
        """
        Validates that the number of columns in the XLS file matches the number of expected fields.
        :param sheet: sheet of the workbook (xlrd.sheet)
        :param expected_fields: list of expected field names
        :param header_row: header row number (default 0)
        :raises ValidationError: if the number of columns does not match
        """
        num_columns = sheet.ncols
        expected_columns = len(expected_fields)

        if num_columns != expected_columns:
            raise ValidationError(
                f"The number of columns ({num_columns}) does not match the expected number ({expected_columns}).\n"
                f"Expected fields: {', '.join(expected_fields)}"
            )
