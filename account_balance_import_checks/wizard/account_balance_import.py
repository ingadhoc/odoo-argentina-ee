import base64
import datetime
import io
import numbers

import xlrd
from odoo import Command, api, fields, models
from odoo.exceptions import ValidationError
from openpyxl import load_workbook


class AccountBalanceImport(models.TransientModel):
    _inherit = "account.balance_import_wizard"

    mode = fields.Selection(
        selection_add=[("check_balance", "Cheques")],
        ondelete={"check_balance": "cascade"},
    )
    check_type = fields.Selection(
        [("issue_check", "Propios"), ("third_check", "De Terceros")],
        string="Tipo de Cheques",
        default="issue_check",
    )

    @api.depends("mode", "check_type")
    def _compute_journal_domain(self):
        checks = self.filtered(lambda r: r.mode == "check_balance")
        for rec in checks:
            if rec.check_type == "issue_check":
                rec.journal_domain = [
                    ("type", "=", "bank"),
                    ("outbound_payment_method_line_ids.code", "=", "own_checks"),
                ]
            else:  # third_check
                rec.journal_domain = [
                    ("type", "=", "cash"),
                    ("inbound_payment_method_line_ids.code", "=", "in_third_party_checks"),
                ]
        super(AccountBalanceImport, self - checks)._compute_journal_domain()

    def action_import(self):
        if self.mode == "check_balance":
            return self._import_checks()
        return super().action_import()

    @api.model
    def locate_bank(self, bank_value):
        """Return res.bank record matching the value from XLS, or empty recordset."""
        bank = self.env["res.bank"]
        if not bank_value:
            return bank
        if isinstance(bank_value, numbers.Number):
            try:
                number_as_float = float(bank_value)
            except (TypeError, ValueError):
                bank_value = str(bank_value)
            else:
                if number_as_float.is_integer():
                    bank_value = str(int(number_as_float))
                else:
                    bank_value = str(number_as_float)
        else:
            bank_value = str(bank_value)
        bank_value = bank_value.strip()
        if not bank_value:
            return bank
        return bank.search([("name", "=", bank_value)], limit=1)

    def _load_check_rows(self, decoded):
        """Load XLS/XLSX bytes and return (rows, workbook_datemode).

        rows is a list of lists with all non-empty rows (including header).
        workbook_datemode is set only for xlrd files; None for openpyxl.
        """
        rows = []
        workbook_datemode = None
        try:
            wb = xlrd.open_workbook(file_contents=decoded)
            sheet = wb.sheet_by_index(0)
            workbook_datemode = wb.datemode
            for row_no in range(sheet.nrows):
                rows.append([cell.value for cell in sheet.row(row_no)])
        except xlrd.biffh.XLRDError:
            wb = load_workbook(io.BytesIO(decoded), data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                rows.append(list(row))
        rows = [row for row in rows if any(cell is not None and cell != "" for cell in row)]
        return rows, workbook_datemode

    def _parse_check_payment_date(self, date_val, workbook_datemode):
        """Return a date object from a cell value (datetime, date, or xlrd float)."""
        if isinstance(date_val, datetime.datetime):
            return date_val.date()
        if isinstance(date_val, datetime.date):
            return date_val
        if workbook_datemode is not None:
            return datetime.datetime(*xlrd.xldate_as_tuple(date_val, workbook_datemode)).date()
        raise TypeError(f"Tipo de fecha no reconocido: {type(date_val)}")

    def _get_check_journal_info(self):
        """Return (journal, payment_method_line) for the current check_type."""
        journal = self.journal_id
        if self.check_type == "issue_check":
            payment_method_line = journal._get_available_payment_method_lines("outbound").filtered(
                lambda x: x.code == "own_checks"
            )
        else:
            payment_method_line = journal._get_available_payment_method_lines("inbound").filtered(
                lambda x: x.code == "new_third_party_checks"
            )
        return journal, payment_method_line

    def _import_checks(self):
        """Triggered on when Checks XLS is imported"""

        # Set-up environment company
        self = self.with_company(self.company_id.id)

        fields = [
            "number",
            "amount",
            "payment_date",
            "name",
            "currency",
            "amount_company_currency",
        ]
        if self.check_type == "third_check":
            fields.insert(4, "owner_vat")
            fields.append("bank_id")

        pre_data = list()  # For storing data before persisting to db
        errors = list()  # For storing possible errors

        decoded = base64.decodebytes(self.file)
        rows, workbook_datemode = self._load_check_rows(decoded)

        # Validate column count
        num_columns = len(rows[0]) if rows else 0
        if num_columns < len(fields):
            raise ValidationError(
                "El archivo importado no tiene el número correcto de columnas. "
                f"Se esperaban {len(fields)} columnas y se encontraron {num_columns}."
            )

        journal, payment_method_line = self._get_check_journal_info()

        # Iterate over each data row (skip header row 0)
        for row_no, row in enumerate(rows[1:], start=1):
            # Create a dict with current row data
            dict_data = {fields[i]: row[i] for i in range(len(fields))}

            # Parse name as string to prevent CUIT being read as float
            if isinstance(dict_data["name"], numbers.Number):
                dict_data["name"] = str(int(dict_data["name"])).strip()
            # Locate Partner
            domain = [
                "|",
                "|",
                ("name", "=", dict_data["name"]),
                ("vat", "=", dict_data["name"]),
                ("ref", "=", dict_data["name"]),
            ]

            partner = self.env["res.partner"].search(domain)

            # Locate Currency
            other_currency = self.locate_currency(dict_data["currency"])

            # Locate Bank (only for third-party checks)
            bank = self.env["res.bank"]
            if self.check_type == "third_check":
                bank = self.locate_bank(dict_data.get("bank_id"))
                if dict_data.get("bank_id") and not bank:
                    errors.append(
                        f"Fila {str(row_no + 1)}: No se encontró un banco para el valor ingresado ({dict_data['bank_id']})."
                    )
                    continue

            if other_currency and not dict_data["amount_company_currency"]:
                errors.append(
                    f"Fila {str(row_no + 1)}: Si le establece otra moneda debe indicar el importe en esa otra moneda"
                )
                continue

            # Skip if partner not found
            if not partner:
                errors.append(
                    "Fila {}: No se encontró ningún partner para el texto ingresado ({})".format(
                        str(row_no + 1), dict_data["name"]
                    )
                )
                continue

            partner = partner.mapped("commercial_partner_id")
            # Skip if more than one partner was found
            if len(partner) > 1:
                errors.append(
                    "Fila {}: Se encontraron varios partners "
                    "para el texto ingresado ({}). ¡Revise los datos Cargados!".format(
                        str(row_no + 1), dict_data["name"]
                    )
                )
                continue

            # Skip if we're not able to parse due_date
            try:
                payment_date = self._parse_check_payment_date(dict_data["payment_date"], workbook_datemode)
            except (TypeError, ValueError):
                errors.append(
                    f"Fila {str(row_no + 1)}: Formato de fecha de pago desconocido. "
                    "Asegúrese de que la columna posee "
                    "formato de fecha."
                )
                continue

            try:
                number = str(int(dict_data["number"]))
            except ValueError:
                errors.append(f"Fila {str(row_no + 1)}: Número de cheque inválido. ")
                continue

            amount_company_currency = (
                other_currency and other_currency.round(dict_data["amount_company_currency"]) or False
            )

            check_data = {
                "partner_id": partner.id,
                "l10n_latam_new_check_ids": [
                    Command.create(
                        {
                            "name": number,
                            "amount": amount_company_currency if amount_company_currency else dict_data["amount"],
                            "bank_id": bank.id if self.check_type == "third_check" and bank else journal.bank_id.id,
                            "payment_date": payment_date,
                        }
                    )
                ],
                "name": f"Saldo Inicial {partner.name} - Cheque Nº. {number}",
                "journal_id": journal.id,
                "payment_type": "outbound" if self.check_type == "issue_check" else "inbound",
                "partner_type": "supplier" if self.check_type == "issue_check" else "customer",
                "date": self.accounting_date,
                "payment_method_line_id": payment_method_line[:1].id,
                "currency_id": other_currency.id if other_currency else self.company_id.currency_id.id,
                "destination_account_id": self.counterpart_account_id.id,
            }

            pre_data.append(check_data)

        # Check if there were errors when iterating over XLS rows
        if len(errors) > 0:
            raise ValidationError("\n".join(errors))

        if not pre_data:
            raise ValidationError("El archivo importado no contiene movimientos.")

        payments = self.env["account.payment"].create(pre_data)
        payments.action_post()

        return payments._get_records_action(name="Cheques Importados", target="current")
